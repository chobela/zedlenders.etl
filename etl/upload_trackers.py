#!/usr/bin/env python3
"""
Jutem GPS Trackers Upload Script
==================================
Reads jutem_fund.xlsx TRACKING FEES sections and creates GPS tracker
records in Directus, assigned to the corresponding borrowers.

Usage:
    python3 etl/upload_trackers.py                # Dry run
    python3 etl/upload_trackers.py --execute      # Actually upload
"""

import os
import sys
import csv
import argparse
import time
import requests
import re
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# ─── Configuration ───────────────────────────────────────────────────
DIRECTUS_URL = "https://zedlenders.pickmesms.com"
COMPANY_ID = 22

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "..", "jutem_fund.xlsx")
NAME_TO_NRC_PATH = os.path.join(os.path.dirname(__file__), "output", "name_to_nrc.csv")

# Section end keywords
SECTION_END_KEYWORDS = [
    "TOTAL", "INTEREST", "COMMISSION", "OPERATIONAL",
    "NET", "PROFIT", "EXPENSE", "INCOME",
]


def get_token():
    token = os.environ.get("DIRECTUS_TOKEN")
    if not token:
        token = input("Enter Directus admin token: ").strip()
    return token


def load_name_to_nrc():
    """Load canonical_name → NRC mapping from ETL output."""
    if not os.path.exists(NAME_TO_NRC_PATH):
        print(f"WARNING: {NAME_TO_NRC_PATH} not found.")
        return {}

    mapping = {}
    with open(NAME_TO_NRC_PATH, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row["canonical_name"].strip()
            mapping[name.upper()] = row["nrc"].strip()

    print(f"Loaded {len(mapping)} name-to-NRC mappings")
    return mapping


def parse_tracking_fees(excel_path):
    """Extract tracking fee rows from all sheets in jutem_fund.xlsx."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    entries = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        in_tracking = False

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            cells = [c.value for c in row]
            first = str(cells[0] or "").strip()
            first_upper = first.upper()

            # Detect TRACKING FEES section header
            if "TRACKING" in first_upper and "FEE" in first_upper:
                in_tracking = True
                continue

            if not in_tracking:
                continue

            # End of section
            if all(c is None for c in cells):
                in_tracking = False
                continue

            if first_upper and any(kw in first_upper for kw in SECTION_END_KEYWORDS):
                in_tracking = False
                continue

            # Skip header row (NAME, DATE, FEE)
            if first_upper in ("NAME", ""):
                continue

            # Parse the entry
            raw_name = first
            raw_date = cells[1] if len(cells) > 1 else None
            raw_fee = cells[2] if len(cells) > 2 else None

            if raw_fee is None or raw_fee == 0:
                continue

            # Extract vehicle plate from name if present
            # Patterns: "Name (PLATE)" or "Tracking Device Installation -PLATE (Name)"
            vehicle_plate = None
            borrower_name = raw_name

            # Pattern: "Tracking Device Installation -BAX 8999ZM (Erica Tembo)"
            install_match = re.match(
                r"Tracking Device Installation\s*[-–]\s*(.+?)\s*\((.+?)\)\s*$",
                raw_name, re.IGNORECASE
            )
            if install_match:
                vehicle_plate = install_match.group(1).strip()
                borrower_name = install_match.group(2).strip()
            else:
                # Pattern: "Mathew Ndaba (BAZ 8998)"
                plate_match = re.match(r"(.+?)\s*\(([A-Z0-9 ]+)\)\s*$", raw_name)
                if plate_match:
                    borrower_name = plate_match.group(1).strip()
                    vehicle_plate = plate_match.group(2).strip()

            # Parse date
            date_issued = None
            if isinstance(raw_date, datetime):
                date_issued = raw_date.strftime("%Y-%m-%d")
            elif raw_date:
                try:
                    date_issued = datetime.strptime(str(raw_date).strip(), "%Y-%m-%d").strftime("%Y-%m-%d")
                except ValueError:
                    pass

            # If no date, derive from sheet name
            if not date_issued:
                try:
                    sheet_dt = datetime.strptime(sheet_name.strip(), "%B %Y")
                    date_issued = sheet_dt.strftime("%Y-%m-01")
                except ValueError:
                    pass

            fee = float(raw_fee) if raw_fee else 0

            entries.append({
                "borrower_name": borrower_name,
                "vehicle_plate": vehicle_plate,
                "date_issued": date_issued,
                "fee": fee,
                "sheet": sheet_name,
            })

    return entries


def generate_dummy_ids(index):
    """Generate dummy GPS code, serial number, and SIM number."""
    gps_code = f"3585110{90000000 + index}"
    serial_number = f"JUTEM-TRK-{index:04d}"
    sim_number = f"+26097800{index:04d}"
    return gps_code, serial_number, sim_number


def lookup_borrowers(token, nrcs):
    """Look up borrower IDs from Directus by NRC numbers."""
    headers = {"Authorization": f"Bearer {token}"}
    nrc_to_id = {}

    batch_size = 20
    nrc_list = list(nrcs)
    for i in range(0, len(nrc_list), batch_size):
        batch = nrc_list[i : i + batch_size]
        nrcs_str = ",".join(batch)
        res = requests.get(
            f"{DIRECTUS_URL}/users",
            headers=headers,
            params={
                "filter[nrc][_in]": nrcs_str,
                "filter[company]": COMPANY_ID,
                "fields": "id,nrc,first_name,last_name",
                "limit": -1,
            },
        )
        res.raise_for_status()
        for user in res.json().get("data", []):
            nrc = user.get("nrc", "")
            if nrc:
                nrc_to_id[nrc] = {
                    "id": user["id"],
                    "name": f"{user.get('first_name', '')} {user.get('last_name', '')}".strip(),
                }

    return nrc_to_id


def lookup_borrower_loans(token, borrower_ids):
    """Look up active loans for borrowers to link trackers."""
    headers = {"Authorization": f"Bearer {token}"}
    borrower_to_loan = {}

    batch_size = 20
    id_list = list(borrower_ids)
    for i in range(0, len(id_list), batch_size):
        batch = id_list[i : i + batch_size]
        ids_str = ",".join(str(bid) for bid in batch)
        res = requests.get(
            f"{DIRECTUS_URL}/items/loans",
            headers=headers,
            params={
                "filter[borrower][_in]": ids_str,
                "filter[company][_eq]": COMPANY_ID,
                "fields": "id,borrower,status",
                "sort": "-date_created",
                "limit": -1,
            },
        )
        res.raise_for_status()
        for loan in res.json().get("data", []):
            borrower_id = loan.get("borrower")
            if borrower_id and borrower_id not in borrower_to_loan:
                # Take the first (most recent) loan for each borrower
                borrower_to_loan[borrower_id] = loan["id"]

    return borrower_to_loan


def check_existing_trackers(token):
    """Check which trackers already exist for this company."""
    headers = {"Authorization": f"Bearer {token}"}
    try:
        res = requests.get(
            f"{DIRECTUS_URL}/items/gps_trackers",
            headers=headers,
            params={
                "filter[company][_eq]": COMPANY_ID,
                "fields": "id,serial_number,issued_to",
                "limit": -1,
            },
        )
        res.raise_for_status()
        existing = res.json().get("data", [])
        return {t["serial_number"]: t for t in existing if t.get("serial_number")}
    except requests.RequestException as e:
        print(f"  WARNING: Could not check existing trackers ({e}). Proceeding without dedup check.")
        return {}


def delete_existing_trackers(token):
    """Delete all trackers for COMPANY_ID."""
    headers = {"Authorization": f"Bearer {token}"}
    print(f"\nDeleting existing trackers for company {COMPANY_ID}...")

    try:
        res = requests.get(
            f"{DIRECTUS_URL}/items/gps_trackers",
            headers=headers,
            params={
                "filter[company][_eq]": COMPANY_ID,
                "fields": "id",
                "limit": -1,
            },
        )
        res.raise_for_status()
    except requests.RequestException as e:
        print(f"  ERROR: Could not fetch existing trackers: {e}")
        return
    items = res.json().get("data", [])

    if not items:
        print("  No existing trackers found.")
        return

    ids = [item["id"] for item in items]
    print(f"  Found {len(ids)} trackers to delete.")

    for item_id in ids:
        res = requests.delete(
            f"{DIRECTUS_URL}/items/gps_trackers/{item_id}",
            headers=headers,
        )
        res.raise_for_status()

    print(f"  Successfully deleted {len(ids)} trackers.")


def main():
    parser = argparse.ArgumentParser(description="Upload Jutem GPS trackers to Directus")
    parser.add_argument("--execute", action="store_true", help="Actually upload (default is dry run)")
    parser.add_argument("--delete-existing", action="store_true",
                        help="Delete all existing trackers before uploading")
    args = parser.parse_args()

    if not os.path.exists(EXCEL_FILE):
        print(f"ERROR: {EXCEL_FILE} not found.")
        sys.exit(1)

    # Parse tracking fees from Excel
    print(f"Parsing tracking fees from {EXCEL_FILE}...")
    entries = parse_tracking_fees(EXCEL_FILE)
    print(f"Found {len(entries)} tracking fee entries")

    if not entries:
        print("No tracking fee data found. Nothing to upload.")
        return

    # Load name-to-NRC mapping
    name_to_nrc = load_name_to_nrc()

    # Resolve borrower NRCs
    for entry in entries:
        name_upper = entry["borrower_name"].upper().strip()
        nrc = name_to_nrc.get(name_upper)
        if not nrc:
            # Try partial match for single-word names like "Muchemwa" or "Greaves"
            for canonical, canonical_nrc in name_to_nrc.items():
                if canonical.startswith(name_upper) or name_upper in canonical:
                    nrc = canonical_nrc
                    break
        entry["nrc"] = nrc

    # Print summary
    print(f"\nTracking fee entries:")
    print(f"{'#':>3}  {'Name':<40} {'Plate':<14} {'Date':<12} {'Fee':>8}  {'NRC':<15}")
    print("-" * 100)
    for i, e in enumerate(entries):
        print(
            f"{i+1:>3}  {e['borrower_name']:<40} {e['vehicle_plate'] or '-':<14} "
            f"{e['date_issued'] or '-':<12} {e['fee']:>8,.0f}  {e['nrc'] or 'NOT FOUND':<15}"
        )

    unresolved = [e for e in entries if not e.get("nrc")]
    if unresolved:
        print(f"\nWARNING: {len(unresolved)} entries have no NRC mapping:")
        for e in unresolved:
            print(f"  - {e['borrower_name']} ({e['sheet']})")

    if not args.execute:
        print("\n--- DRY RUN ---")
        print("Run with --execute to actually upload.")
        print("Run with --execute --delete-existing to clear and re-upload.")
        return

    token = get_token()

    # Delete existing if requested
    if args.delete_existing:
        delete_existing_trackers(token)

    # Look up borrower IDs in Directus
    nrcs_to_lookup = {e["nrc"] for e in entries if e.get("nrc")}
    print(f"\nLooking up {len(nrcs_to_lookup)} borrowers in Directus...")
    nrc_to_borrower = lookup_borrowers(token, nrcs_to_lookup)
    print(f"  Found {len(nrc_to_borrower)} borrowers")

    # Look up loans for those borrowers
    borrower_ids = {b["id"] for b in nrc_to_borrower.values()}
    print(f"Looking up loans for {len(borrower_ids)} borrowers...")
    borrower_to_loan = lookup_borrower_loans(token, borrower_ids)
    print(f"  Found loans for {len(borrower_to_loan)} borrowers")

    # Check existing trackers
    existing = check_existing_trackers(token)
    print(f"Existing trackers: {len(existing)}")

    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }

    successes = 0
    errors = []

    for i, entry in enumerate(entries):
        gps_code, serial_number, sim_number = generate_dummy_ids(i + 1)

        # Skip if serial number already exists
        if serial_number in existing:
            print(f"  [{i+1}/{len(entries)}] SKIP (exists): {serial_number}")
            continue

        # Resolve borrower and loan
        borrower_info = None
        loan_id = None
        if entry.get("nrc"):
            borrower_info = nrc_to_borrower.get(entry["nrc"])
            if borrower_info:
                loan_id = borrower_to_loan.get(borrower_info["id"])

        payload = {
            "gps_code": gps_code,
            "serial_number": serial_number,
            "sim_number": sim_number,
            "gps_fee": entry["fee"],
            "status": "active" if borrower_info else "not_assigned",
            "company": COMPANY_ID,
        }

        if borrower_info:
            payload["issued_to"] = borrower_info["id"]
            payload["date_issued"] = entry["date_issued"]

        if loan_id:
            payload["assigned_loan"] = loan_id

        try:
            res = requests.post(
                f"{DIRECTUS_URL}/items/gps_trackers",
                headers=headers,
                json=payload,
            )
            res.raise_for_status()
            successes += 1
            assigned_to = borrower_info["name"] if borrower_info else "UNASSIGNED"
            loan_str = f" (Loan #{loan_id})" if loan_id else ""
            print(
                f"  [{i+1}/{len(entries)}] Created: {serial_number} → "
                f"{entry['borrower_name']} → {assigned_to}{loan_str} | K {entry['fee']:,.0f}"
            )
        except requests.RequestException as e:
            err_msg = str(e)
            try:
                err_msg = res.json().get("errors", [{}])[0].get("message", err_msg)
            except Exception:
                pass
            errors.append({
                "name": entry["borrower_name"],
                "serial": serial_number,
                "error": err_msg,
            })
            print(f"  [{i+1}/{len(entries)}] FAILED: {serial_number} - {err_msg}")

        time.sleep(0.3)

    print(f"\n{'='*50}")
    print("Upload complete!")
    print(f"  Created: {successes}")
    print(f"  Failed: {len(errors)}")

    if errors:
        for err in errors:
            print(f"  ERROR: {err['name']} ({err['serial']}): {err['error']}")


if __name__ == "__main__":
    main()
