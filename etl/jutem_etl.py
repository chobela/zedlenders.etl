#!/usr/bin/env python3
"""
Jutem Fund ETL Script
=====================
Extracts borrowers and loans from jutem_fund.xlsx and generates
CSV files ready for upload to ZedLenders (Directus).

Company ID: 22

Usage:
    python3 etl/jutem_etl.py                  # Run full ETL
    python3 etl/jutem_etl.py --step borrowers # Generate borrowers CSV only
    python3 etl/jutem_etl.py --step loans     # Generate loans CSV only
    python3 etl/jutem_etl.py --step review    # Generate review CSV for name dedup
"""

import os
import re
import csv
import sys
import argparse
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# ─── Configuration ───────────────────────────────────────────────────────────

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "..", "jutem_fund.xlsx")
LIABILITIES_FILE = os.path.join(os.path.dirname(__file__), "..", "liabilities.xlsx")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "output")
COMPANY_ID = 22

# Placeholder sequences
NRC_START = 1000001
PHONE_START = 260977000001
EMAIL_DOMAIN = "@kwachaplus.com"

# Exact names to skip (non-loan section headers / summary rows)
SKIP_EXACT = {
    "TOTAL", "NAME", "DATE", "COMMISSION", "BORROWED",
    "INTEREST INCOME", "INTEREST EXPENSE", "NET INTEREST INCOME",
    "NET OPERATING INCOME", "OTHER EXPENSES", "TOTAL EXPENSES",
    "TOTAL INCOME", "NET INCOME",
    "NET INTEREST INCOME AND TRACKING FEE",
}

# Keywords that indicate a non-loan row ONLY when the entire name is an expense label
# (not when the keyword appears inside a person's name like "Theresa (Money Acumen)")
SKIP_KEYWORDS = [
    "EXPENSE", "INCOME", "DEPRECIATION", "SALARY",
    "OPERATING", "LABOUR", "LABOR", "MAINTENANCE", "INSURANCE",
    "STATIONERY", "BAD DEBT", "TRACKING FEE",
]

# Known name aliases → canonical name mapping
# This handles the most obvious duplicates found in the data.
# After running --step review, you can add more mappings here.
NAME_ALIASES = {
    # Parenthetical references - strip the parenthetical
    "Agness(Barry)": "Agness",
    "Alex(claudius)": "Alex",
    "Albert(jastow)": "Albert Jastow",
    "Albert(Jastrow)": "Albert Jastow",
    "Albert jast": "Albert Jastow",
    "Kantu(Albert)": "Kantu",
    "Mwenya(Dudu)": "Mwenya",
    "Michale(Ruth)": "Michale",
    "Newton(Ruth)": "Newton",
    "Nelson(Rocky)": "Nelson",
    "Ruth(Chikunta)": "Ruth Chikunta",
    "RUTH CHIKUNTA": "Ruth Chikunta",
    "Kingsley(IDC)": "Kingsley",
    "Charles IDC": "Charles IDC",
    "Anthony IDC": "Anthony IDC",
    "Brenda(likito)": "Brenda Likito",
    "A.B.M (Kwezekani)": "ABM Kwezekani",
    "Benson (Mwaba)": "Benson Mwaba",

    # Spelling variations
    "Ba Chrisitabel": "Ba Christabel",
    "Ba Chrsitabel": "Ba Christabel",
    "Abigail Beenzu": "Abigail Benzu",
    "Abigial": "Abigail",
    "Abiagail": "Abigail",
    "Annastasia": "Anastasia",
    "Annastsia": "Anastasia",
    "Anastasi": "Anastasia",
    "Aunty jane": "Aunt Jane",
    "Aunty Jane": "Aunt Jane",
    "Aunt jane": "Aunt Jane",
    "Amingtone": "Amington",
    "Amington": "Amington",
    "Auther": "Aurther",
    "Benadette": "Bernadette",
    "Benaddete": "Bernadette",
    "benadete": "Bernadette",
    "Benadettw": "Bernadette",
    "Barbra Chikondo": "Barbara Chikondo",
    "christabel": "Christabel",
    "cleopatra": "Cleopatra",
    "Kindele": "Kindile",
    "Mwansa Chulo": "Mwansa Chalo",
    "A.milimo": "Milimo",
    "A.Milimo": "Milimo",

    # Mr/Mrs prefixes - keep as-is but normalize
    "Mr. Mwale": "Mr Mwale",
    "Mr.Yombo": "Mr Yombo",
    "Mr.Tonga": "Mr Tonga",
    "Mr Tonga ": "Mr Tonga",
    "Mr. Tonga": "Mr Tonga",
    "Mr. Valand": "Mr Valand",
    "Mr. Mukusayi": "Mr Mukusayi",
    "Mr Bridgetis": "Mr Bridgetis",

    # Br prefix
    "Br Bwali": "Br Bwali",
    "Br Kalinda": "Br Kalinda",
    "Br Mumba": "Br Mumba",
    "Br. Mumba(Elderly)": "Br Mumba",
    "Br.Mumba": "Br Mumba",

    # Case/spacing
    "G.B": "GB",
    "KONDWA": "Kondwa",
    "Uncle chris ": "Uncle Chris",
    "Uncle chris": "Uncle Chris",
}


# ─── Helper Functions ────────────────────────────────────────────────────────

def parse_sheet_date(sheet_name):
    """Convert sheet name like 'MAY 2021' to (year, month) tuple."""
    try:
        dt = datetime.strptime(sheet_name.strip(), "%B %Y")
        return dt.year, dt.month
    except ValueError:
        return None, None


def parse_day_from_text(text, year, month):
    """
    Parse messy date text like '20th may', '1st', '2nd', '15TH Oct' into a date.
    Returns ISO date string or None.
    """
    if not text or not isinstance(text, str):
        return None

    text = text.strip()

    # Extract day number
    day_match = re.search(r"(\d{1,2})", text)
    if not day_match:
        return None

    day = int(day_match.group(1))
    if day < 1 or day > 31:
        return None

    # Clamp day to valid range for the month
    import calendar
    max_day = calendar.monthrange(year, month)[1]
    day = min(day, max_day)

    try:
        return f"{year}-{month:02d}-{day:02d}"
    except ValueError:
        return f"{year}-{month:02d}-01"


def clean_name(raw_name):
    """Clean and normalize a borrower name."""
    if not raw_name or not isinstance(raw_name, str):
        return None

    name = raw_name.strip()

    # Check alias mapping first
    if name in NAME_ALIASES:
        name = NAME_ALIASES[name]

    # Title case
    name = name.strip()

    # Remove trailing/leading whitespace
    name = re.sub(r"\s+", " ", name)

    return name


def split_first_last(canonical_name):
    """
    Split a canonical name into (first_name, last_name).
    - Single word → (word, "N/A")
    - Two+ words → (first, rest)
    """
    parts = canonical_name.strip().split()
    if len(parts) == 0:
        return ("Unknown", "N/A")
    elif len(parts) == 1:
        return (parts[0], "N/A")
    else:
        return (parts[0], " ".join(parts[1:]))


def is_loan_row(name_str):
    """Check if a row is a loan entry (not a summary/expense/header row)."""
    if not name_str or not isinstance(name_str, str):
        return False

    stripped = name_str.strip()
    upper = stripped.upper()

    # Skip exact matches (section headers, summary rows)
    if upper in SKIP_EXACT:
        return False

    # Skip rows where the entire name is an expense label
    for keyword in SKIP_KEYWORDS:
        if keyword in upper:
            return False

    # Skip very short strings
    if len(stripped) <= 1:
        return False

    # Skip date-like entries
    if re.match(r"^\d{1,2}/\d{1,2}/\d{2,4}$", stripped):
        return False

    return True


def extract_loans_from_sheet(ws, sheet_name):
    """
    Extract loan rows from a worksheet.
    Only reads rows between the first header and the first TOTAL row.
    """
    year, month = parse_sheet_date(sheet_name)
    if year is None:
        return []

    loans = []
    in_loan_section = False

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # Detect the loan section header
        if i == 0:
            first_cell = str(row[0]).strip().upper() if row[0] else ""
            if first_cell == "NAME":
                in_loan_section = True
                continue

        if not in_loan_section:
            continue

        name_cell = row[0]

        # Stop at TOTAL row (end of loan section)
        if name_cell and isinstance(name_cell, str) and name_cell.strip().upper() == "TOTAL":
            break

        # Skip non-loan rows
        if not is_loan_row(name_cell if isinstance(name_cell, str) else None):
            continue

        # Extract fields
        raw_name = str(name_cell).strip() if name_cell else None
        if not raw_name:
            continue

        amount = row[1] if len(row) > 1 else None
        interest = row[2] if len(row) > 2 else None
        date_text = row[4] if len(row) > 4 else None
        status = row[6] if len(row) > 6 else None

        # Clean amount
        if isinstance(amount, str):
            # Remove formulas
            if amount.startswith("="):
                amount = None
            else:
                try:
                    amount = float(amount.replace(",", ""))
                except ValueError:
                    amount = None

        # Clean interest (stored as decimal like 0.25 = 25%)
        if isinstance(interest, str):
            if interest.startswith("="):
                interest = None
            else:
                try:
                    interest = float(interest)
                except ValueError:
                    interest = None

        if interest and isinstance(interest, (int, float)):
            # Convert decimal to percentage (0.25 → 25)
            interest_pct = round(interest * 100, 2)
        else:
            interest_pct = None

        # Parse date
        date_str = parse_day_from_text(str(date_text) if date_text else None, year, month)
        if not date_str:
            # Default to 1st of the month
            date_str = f"{year}-{month:02d}-01"

        # Clean status
        status_str = str(status).strip().lower() if status else "pending"
        # Map to system status codes
        if status_str in ("paid", "cleared"):
            loan_status = 6  # Closed/Paid
        elif status_str in ("pending", "unsettled", "not cleared"):
            loan_status = 5  # Active/Disbursed
        elif status_str == "moved":
            loan_status = 6  # Treat as closed
        elif status_str == "non":
            loan_status = 5  # Active
        else:
            loan_status = 5  # Default active

        canonical = clean_name(raw_name)
        if not canonical:
            continue

        loans.append({
            "sheet": sheet_name,
            "raw_name": raw_name,
            "canonical_name": canonical,
            "amount": round(amount, 2) if amount else 0,
            "interest_pct": interest_pct,
            "date_expected": date_str,
            "date_expected_raw": str(date_text).strip() if date_text else "",
            "application_date": f"{year}-{month:02d}-01",
            "loan_status": loan_status,
            "status_text": status_str,
        })

    return loans


# ─── Main ETL Steps ──────────────────────────────────────────────────────────

def load_workbook():
    """Load the Excel workbook."""
    path = os.path.abspath(EXCEL_FILE)
    print(f"Loading workbook: {path}")
    wb = openpyxl.load_workbook(path, data_only=False)
    print(f"Found {len(wb.sheetnames)} sheets: {wb.sheetnames[0]} → {wb.sheetnames[-1]}")
    return wb


def extract_all_loans(wb):
    """Extract all loan rows from all sheets."""
    all_loans = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        loans = extract_loans_from_sheet(ws, sheet_name)
        all_loans.extend(loans)
        if loans:
            print(f"  {sheet_name}: {len(loans)} loans")
    print(f"\nTotal loans extracted: {len(all_loans)}")
    return all_loans


def build_borrower_registry(all_loans):
    """
    Build a deduplicated borrower registry from all loans.
    Returns dict: canonical_name → borrower info
    """
    registry = {}
    name_occurrences = defaultdict(int)

    for loan in all_loans:
        canonical = loan["canonical_name"]
        name_occurrences[canonical] += 1

        if canonical not in registry:
            first, last = split_first_last(canonical)
            registry[canonical] = {
                "canonical_name": canonical,
                "first_name": first,
                "last_name": last,
                "loan_count": 0,
                "first_seen": loan["sheet"],
                "last_seen": loan["sheet"],
            }

        registry[canonical]["loan_count"] = name_occurrences[canonical]
        registry[canonical]["last_seen"] = loan["sheet"]

    # Sort by canonical name
    sorted_registry = dict(sorted(registry.items(), key=lambda x: x[0].lower()))
    print(f"Unique borrowers: {len(sorted_registry)}")
    return sorted_registry


def generate_review_csv(registry):
    """Generate a review CSV for manual name deduplication."""
    output_path = os.path.join(OUTPUT_DIR, "borrowers_review.csv")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    with open(output_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "canonical_name", "first_name", "last_name",
            "loan_count", "first_seen", "last_seen", "action"
        ])
        for name, info in registry.items():
            writer.writerow([
                info["canonical_name"],
                info["first_name"],
                info["last_name"],
                info["loan_count"],
                info["first_seen"],
                info["last_seen"],
                "",  # action column for manual review
            ])

    print(f"\nReview CSV written to: {output_path}")
    print(f"  {len(registry)} unique borrowers")
    print("  Review the file, then re-run with --step borrowers")
    return output_path


def generate_borrowers_csv(registry):
    """Generate borrowers upload CSV with placeholder fields."""
    output_path = os.path.join(OUTPUT_DIR, "borrowers_upload.csv")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Build NRC → canonical mapping for loans step
    nrc_map = {}

    with open(output_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "first_name", "last_name", "email", "phone", "nrc",
            "address", "province", "employment_type", "business",
            "reference_number", "department",
            "next_kin_name", "next_kin_phone", "next_kin_relationship"
        ])

        for idx, (canonical, info) in enumerate(registry.items()):
            nrc_num = NRC_START + idx
            phone_num = PHONE_START + idx
            nrc = f"{nrc_num:07d}/00/1"
            phone = str(phone_num)
            email = f"{nrc_num}{EMAIL_DOMAIN}"

            nrc_map[canonical] = nrc

            writer.writerow([
                info["first_name"],
                info["last_name"],
                email,
                phone,
                nrc,
                "Lusaka",            # address placeholder
                "Lusaka",            # province
                "self_employed",     # employment_type
                "N/A",               # business
                str(nrc_num),        # reference_number
                "N/A",               # department
                "N/A",               # next_kin_name
                "260977000000",      # next_kin_phone
                "N/A",               # next_kin_relationship
            ])

    # Save NRC mapping for loans step
    mapping_path = os.path.join(OUTPUT_DIR, "name_to_nrc.csv")
    with open(mapping_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["canonical_name", "nrc"])
        for name, nrc in nrc_map.items():
            writer.writerow([name, nrc])

    print(f"\nBorrowers CSV written to: {output_path}")
    print(f"  {len(registry)} borrowers")
    print(f"NRC mapping saved to: {mapping_path}")
    return output_path, nrc_map


def extract_day_number(date_text):
    """Extract just the day number from date text for grouping purposes."""
    if not date_text:
        return None

    text = str(date_text).strip()

    # Handle datetime strings like "2024-10-25 00:00:00"
    dt_match = re.match(r"\d{4}-\d{2}-(\d{2})", text)
    if dt_match:
        return int(dt_match.group(1))

    # Handle text like "25th", "1st", "2nd"
    match = re.search(r"(\d{1,2})", text)
    if match:
        day = int(match.group(1))
        if 1 <= day <= 31:
            return day
    return None


def group_loans(all_loans):
    """
    Group installment rows into actual loans.

    Grouping key: (canonical_name, monthly_amount_rounded, day_of_month)
    - Same borrower + same monthly amount + same expected day = one loan
    - Total principal = monthly_amount × number_of_months
    - Application date = first month the installment appears
    - Status = Settled (6) only if ALL installments are Paid, otherwise Active (5)
    """
    # First pass: extract day numbers for each installment row
    installments = []
    for loan in all_loans:
        if not loan["amount"] or loan["amount"] <= 0:
            continue

        amount_key = round(loan["amount"], 2)

        day = extract_day_number(loan.get("date_expected_raw", ""))
        if day is None:
            date_str = loan.get("date_expected", "")
            if date_str and len(date_str) == 10:
                try:
                    day = int(date_str.split("-")[2])
                except (IndexError, ValueError):
                    day = 1
            else:
                day = 1

        installments.append((loan, amount_key, day))

    # Second pass: group with ±1 day tolerance
    # For each (name, amount) pair, cluster days that are within 1 of each other
    # e.g. days [24, 25, 25, 25] → all become day 25 (the mode)
    from collections import Counter

    # Collect all days per (name, amount) to find the dominant day
    name_amount_days = defaultdict(list)
    for loan, amount_key, day in installments:
        name_amount_days[(loan["canonical_name"], amount_key)].append(day)

    # For each (name, amount), find day clusters and pick the mode for each cluster
    day_mapping = {}  # (name, amount, original_day) → canonical_day
    for (name, amount), days in name_amount_days.items():
        # Sort unique days and cluster those within ±1
        unique_days = sorted(set(days))
        clusters = []
        current_cluster = [unique_days[0]]

        for d in unique_days[1:]:
            if d - current_cluster[-1] <= 1:
                current_cluster.append(d)
            else:
                clusters.append(current_cluster)
                current_cluster = [d]
        clusters.append(current_cluster)

        # For each cluster, find the most common day (mode)
        day_counts = Counter(days)
        for cluster in clusters:
            mode_day = max(cluster, key=lambda d: day_counts[d])
            for d in cluster:
                day_mapping[(name, amount, d)] = mode_day

    # Build groups using the canonical day
    groups = defaultdict(list)
    for loan, amount_key, day in installments:
        canonical_day = day_mapping.get(
            (loan["canonical_name"], amount_key, day), day
        )
        key = (loan["canonical_name"], amount_key, canonical_day)
        groups[key].append(loan)

    # Convert groups to loan records
    grouped_loans = []
    for (name, monthly_amount, day), installments in groups.items():
        # Sort by application_date to get chronological order
        installments.sort(key=lambda x: x["application_date"])

        num_months = len(installments)
        total_amount = round(monthly_amount * num_months, 2)
        application_date = installments[0]["application_date"]
        first_sheet = installments[0]["sheet"]
        last_sheet = installments[-1]["sheet"]

        # Interest rate from the first installment
        interest_pct = installments[0].get("interest_pct") or 25

        # Status: settled ONLY if ALL installments are paid
        all_paid = all(inst["status_text"] in ("paid", "cleared", "moved") for inst in installments)
        loan_status = 6 if all_paid else 5  # 6=Settled, 5=Active

        grouped_loans.append({
            "canonical_name": name,
            "total_amount": total_amount,
            "monthly_amount": monthly_amount,
            "num_months": num_months,
            "interest_pct": interest_pct,
            "application_date": application_date,
            "loan_status": loan_status,
            "day_expected": day,
            "first_sheet": first_sheet,
            "last_sheet": last_sheet,
            "installments": installments,  # keep raw rows for amortization
        })

    # Sort by borrower name then application date
    grouped_loans.sort(key=lambda x: (x["canonical_name"].lower(), x["application_date"]))

    print(f"\nGrouped {len(all_loans)} installment rows → {len(grouped_loans)} actual loans")

    # Print some stats
    settled = sum(1 for l in grouped_loans if l["loan_status"] == 6)
    active = sum(1 for l in grouped_loans if l["loan_status"] == 5)
    multi = sum(1 for l in grouped_loans if l["num_months"] > 1)
    single = sum(1 for l in grouped_loans if l["num_months"] == 1)
    print(f"  Settled: {settled}, Active: {active}")
    print(f"  Multi-month loans: {multi}, Single-month loans: {single}")

    return grouped_loans


def generate_loans_csv(all_loans, nrc_map):
    """Generate loans upload CSV and amortization CSV."""
    loans_path = os.path.join(OUTPUT_DIR, "loans_upload.csv")
    amort_path = os.path.join(OUTPUT_DIR, "amortization_upload.csv")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    grouped = group_loans(all_loans)

    missing_nrc = set()
    loans_written = 0
    amort_written = 0

    loans_file = open(loans_path, "w", newline="")
    amort_file = open(amort_path, "w", newline="")

    loans_writer = csv.writer(loans_file)
    amort_writer = csv.writer(amort_file)

    loans_writer.writerow([
        "loan_reference", "borrower_nrc", "loan_product_name", "amount",
        "custom_interest_rate", "application_date", "loan_purpose",
        "branch_name", "loan_status", "approval_date",
        "amount_type", "monthly_amount"
    ])

    amort_writer.writerow([
        "loan_reference", "borrower_nrc", "month", "due_date",
        "amount_due", "interest_rate", "expected_amount",
        "profit", "status", "sheet"
    ])

    for loan_idx, loan in enumerate(grouped, start=1):
        canonical = loan["canonical_name"]
        nrc = nrc_map.get(canonical)

        if not nrc:
            missing_nrc.add(canonical)
            continue

        # Loan reference: JUTEM-{sequential_id}
        loan_ref = f"JUTEM-{loan_idx:05d}"

        # Write grouped loan
        loans_writer.writerow([
            loan_ref,
            nrc,
            "Business Micro Loan",
            loan["total_amount"],
            loan["interest_pct"],
            loan["application_date"],
            "Business",
            "Lusaka",
            loan["loan_status"],
            loan["application_date"],
            "",
            round(loan["monthly_amount"], 2),
        ])
        loans_written += 1

        # Write each installment as an amortization row
        for inst_idx, inst in enumerate(loan["installments"], start=1):
            year, month = parse_sheet_date(inst["sheet"])
            day = loan["day_expected"]
            if year and month:
                import calendar
                max_day = calendar.monthrange(year, month)[1]
                day = min(day, max_day)
                due_date = f"{year}-{month:02d}-{day:02d}"
            else:
                due_date = inst["application_date"]

            # Calculate expected amount: amount + (amount * interest)
            amount_due = inst["amount"]
            interest_decimal = (inst["interest_pct"] or 0) / 100.0
            expected_amount = round(amount_due * (1 + interest_decimal), 2)
            profit = round(amount_due * interest_decimal, 2)

            amort_writer.writerow([
                loan_ref,
                nrc,
                inst["sheet"],
                due_date,
                round(amount_due, 2),
                inst["interest_pct"],
                expected_amount,
                profit,
                inst["status_text"],
                inst["sheet"],
            ])
            amort_written += 1

    loans_file.close()
    amort_file.close()

    if missing_nrc:
        print(f"\nWARNING: {len(missing_nrc)} borrowers had no NRC mapping:")
        for name in sorted(missing_nrc):
            print(f"  - {name}")

    print(f"\nLoans CSV written to: {loans_path}")
    print(f"  {loans_written} loans written")
    print(f"\nAmortization CSV written to: {amort_path}")
    print(f"  {amort_written} installment rows written")

    # Also write a detailed grouped loans report
    report_path = os.path.join(OUTPUT_DIR, "loans_grouped_report.csv")
    with open(report_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "loan_reference", "borrower", "monthly_amount", "num_months",
            "total_amount", "interest_pct", "day_expected", "application_date",
            "status", "first_sheet", "last_sheet"
        ])
        for loan_idx, loan in enumerate(grouped, start=1):
            writer.writerow([
                f"JUTEM-{loan_idx:05d}",
                loan["canonical_name"],
                loan["monthly_amount"],
                loan["num_months"],
                loan["total_amount"],
                loan["interest_pct"],
                loan["day_expected"],
                loan["application_date"],
                "Settled" if loan["loan_status"] == 6 else "Active",
                loan["first_sheet"],
                loan["last_sheet"],
            ])
    print(f"Grouped report written to: {report_path}")


# ─── Step 4: Payments ────────────────────────────────────────────────────────

def generate_payments_csv():
    """
    Generate payments_upload.csv from the amortization schedule.

    For every amortization row with status 'paid' or 'cleared', create a
    payment record whose amount equals the expected_amount (principal + interest)
    from that amortization entry.  The payment date is set to the amortization
    due_date.

    Output format matches public/sample_payments_upload.csv:
        loan_reference,payment_amount,payment_date,payment_method,transaction_type,reference_number
    """
    amort_path = os.path.join(OUTPUT_DIR, "amortization_upload.csv")
    payments_path = os.path.join(OUTPUT_DIR, "payments_upload.csv")

    if not os.path.exists(amort_path):
        print(f"ERROR: {amort_path} not found. Run --step loans first.")
        sys.exit(1)

    # Read amortization rows
    amort_rows = []
    with open(amort_path, "r") as f:
        reader = csv.DictReader(f)
        for row in reader:
            amort_rows.append(row)

    print(f"\nGenerating payments from {len(amort_rows)} amortization rows...")

    payments_written = 0
    pmt_counter = 1

    with open(payments_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "loan_reference", "payment_amount", "payment_date",
            "payment_method", "transaction_type", "reference_number"
        ])

        for row in amort_rows:
            status = row.get("status", "").strip().lower()

            # Only generate payments for paid/cleared installments
            if status not in ("paid", "cleared"):
                continue

            loan_ref = row["loan_reference"]
            expected_amount = float(row["expected_amount"])
            due_date = row["due_date"]
            ref_number = f"PMT-{pmt_counter:05d}"

            writer.writerow([
                loan_ref,
                expected_amount,
                due_date,
                "cash",
                "Loan Repayment",
                ref_number,
            ])

            payments_written += 1
            pmt_counter += 1

    paid_count = sum(1 for r in amort_rows if r.get("status", "").strip().lower() in ("paid", "cleared"))
    skipped = len(amort_rows) - paid_count

    print(f"\nPayments CSV written to: {payments_path}")
    print(f"  {payments_written} payments generated (from paid/cleared installments)")
    print(f"  {skipped} installments skipped (pending/unsettled/non/moved)")


# ─── Step 5: Generate Expenses CSV ──────────────────────────────────────────

def generate_expenses_csv(wb):
    """
    Extract OTHER EXPENSES tables from all sheets and write expenses_upload.csv.

    Each sheet has an 'OTHER EXPENSES' header row followed by:
        DATE | EXPENSE | AMOUNT
    rows until a 'TOTAL' row is hit.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    expenses_path = os.path.join(OUTPUT_DIR, "expenses_upload.csv")

    fieldnames = [
        "expense_description",
        "amount",
        "transaction_date",
        "reference_number",
        "sheet",
    ]

    all_expenses = []
    ref_counter = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        year, month = parse_sheet_date(sheet_name)
        if year is None:
            continue

        # Find the "OTHER EXPENSES" header row
        expense_start = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=1, values_only=False):
            cell_val = row[0].value
            if cell_val and str(cell_val).strip().upper() == "OTHER EXPENSES":
                expense_start = row[0].row
                break

        if expense_start is None:
            continue

        # Skip the header row (DATE | EXPENSE | AMOUNT)
        data_start = expense_start + 2

        for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row, max_col=3, values_only=False):
            vals = [c.value for c in row]
            date_val, description, amount = vals[0], vals[1], vals[2]

            # Stop at TOTAL row or empty row
            if description is None and amount is None:
                break
            if date_val and str(date_val).strip().upper() == "TOTAL":
                break

            # Skip rows with no description or amount
            if not description or not amount:
                continue

            try:
                amount = float(amount)
            except (ValueError, TypeError):
                continue

            if amount <= 0:
                continue

            # Parse the date
            txn_date = None
            if isinstance(date_val, datetime):
                txn_date = date_val.strftime("%Y-%m-%d")
            elif date_val and str(date_val).strip() not in ("-", ""):
                # Try to parse text dates
                try:
                    txn_date = datetime.strptime(str(date_val).strip(), "%Y-%m-%d").strftime("%Y-%m-%d")
                except ValueError:
                    pass

            # Default to 1st of the sheet's month if no valid date
            if not txn_date:
                txn_date = f"{year}-{month:02d}-01"

            ref_counter += 1
            all_expenses.append({
                "expense_description": str(description).strip(),
                "amount": amount,
                "transaction_date": txn_date,
                "reference_number": f"EXP-{ref_counter:05d}",
                "sheet": sheet_name,
            })

    # Write CSV
    with open(expenses_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_expenses)

    print(f"\nExpenses CSV written to: {expenses_path}")
    print(f"  {len(all_expenses)} expense rows extracted from {len(wb.sheetnames)} sheets")

    # Summary by category
    categories = defaultdict(float)
    for e in all_expenses:
        categories[e["expense_description"]] += e["amount"]
    print(f"  Top expense categories:")
    for cat, total in sorted(categories.items(), key=lambda x: -x[1])[:10]:
        print(f"    {cat}: K {total:,.2f}")


# ─── Investors & Investments ETL ──────────────────────────────────────────────

MONTH_COLS = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}


def load_liabilities():
    """Load liabilities.xlsx workbook."""
    if not os.path.exists(LIABILITIES_FILE):
        print(f"ERROR: {LIABILITIES_FILE} not found.")
        sys.exit(1)
    return openpyxl.load_workbook(LIABILITIES_FILE, data_only=True)


def parse_liabilities():
    """
    Parse liabilities.xlsx into a list of investment records.

    The sheet has year sections (2022, 2023, 2024, 2025, 2026) each with:
        NAME | B/F | JAN | FEB | ... | DEC | TOTAL | STATUS (2025+ only)

    Returns list of dicts with: investor_name, principal_amount, year,
    investment_date, status, monthly_amounts
    """
    wb = load_liabilities()
    ws = wb[wb.sheetnames[0]]

    investments = []
    current_year = None
    header_row = None
    has_status = False

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        cells = list(row)
        first_val = cells[0].value

        # Detect year marker (standalone integer in column A)
        if first_val and isinstance(first_val, (int, float)) and 2020 <= first_val <= 2030:
            current_year = int(first_val)
            header_row = None
            continue

        # Detect header row (NAME | B/F | JAN | ...)
        if first_val and str(first_val).strip().upper() == "NAME":
            header_row = cells[0].row
            # Check if STATUS column exists (col P = index 15)
            has_status = len(cells) > 15 and cells[15].value and str(cells[15].value).strip().upper() == "STATUS"
            continue

        if header_row is None or current_year is None:
            continue

        # Skip non-data rows
        if not first_val or str(first_val).strip().upper() in ("TOTAL", "AMOUNT BORROWED", ""):
            if str(first_val or "").strip().upper() == "TOTAL":
                header_row = None  # End of this year section
            continue

        name = str(first_val).strip()
        if not name:
            continue

        # Extract monthly amounts (columns C-N = indices 2-13)
        # B/F is column B = index 1
        bf_val = cells[1].value
        bf_amount = float(bf_val) if bf_val and isinstance(bf_val, (int, float)) else 0

        monthly_amounts = {}
        first_month = None
        for col_name, month_num in MONTH_COLS.items():
            col_idx = month_num + 1  # JAN=C=index 2, FEB=D=index 3, etc.
            if col_idx < len(cells):
                val = cells[col_idx].value
                if val and isinstance(val, (int, float)) and val != 0:
                    monthly_amounts[month_num] = float(val)
                    if first_month is None:
                        first_month = month_num

        # Total column (O = index 14)
        total_val = cells[14].value if len(cells) > 14 else None
        if total_val and isinstance(total_val, (int, float)):
            principal = float(total_val)
        else:
            principal = bf_amount + sum(monthly_amounts.values())

        if principal == 0 and bf_amount == 0:
            continue

        # Status — map to values the frontend expects:
        #   Active (current), Matured (past/completed), Settled (paid off)
        STATUS_MAP = {"Ongoing": "Active", "Paid": "Settled", "Matured": "Matured"}
        raw_status = "Matured"
        if has_status and len(cells) > 15 and cells[15].value:
            raw_status = str(cells[15].value).strip()
        elif current_year >= 2025:
            raw_status = "Ongoing"
        status = STATUS_MAP.get(raw_status, raw_status)

        # Investment date = first month with a value, or Jan 1 if only B/F
        if first_month:
            inv_date = f"{current_year}-{first_month:02d}-01"
        else:
            inv_date = f"{current_year}-01-01"

        investments.append({
            "investor_name": name,
            "principal_amount": round(principal, 2),
            "investment_date": inv_date,
            "year": current_year,
            "status": status,
            "bf_amount": round(bf_amount, 2),
        })

    wb.close()
    return investments


def generate_investors_csv():
    """Extract unique investor names from liabilities.xlsx and write investors_upload.csv."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    investments = parse_liabilities()

    # Collect unique names (case-normalized)
    seen = {}
    for inv in investments:
        key = inv["investor_name"].upper().strip()
        if key not in seen:
            seen[key] = inv["investor_name"]

    investors_path = os.path.join(OUTPUT_DIR, "investors_upload.csv")
    fieldnames = ["name", "status", "company"]

    rows = []
    for key in sorted(seen.keys()):
        rows.append({
            "name": seen[key],
            "status": "Active",
            "company": COMPANY_ID,
        })

    with open(investors_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"\nInvestors CSV written to: {investors_path}")
    print(f"  {len(rows)} unique investors")
    for r in rows:
        print(f"    - {r['name']}")


def generate_investments_csv():
    """
    Extract investment records from liabilities.xlsx.
    One investment per lender per year.
    Output: investments_upload.csv
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    investments = parse_liabilities()

    inv_path = os.path.join(OUTPUT_DIR, "investments_upload.csv")
    fieldnames = [
        "reference_number", "investor_name", "principal_amount",
        "investment_date", "year", "status", "bf_amount",
    ]

    ref_counter = 0
    rows = []
    for inv in investments:
        ref_counter += 1
        safe_name = re.sub(r'[^A-Z0-9]', '', inv["investor_name"].upper())[:10]
        rows.append({
            "reference_number": f"JINV-{safe_name}-{inv['year']}-{ref_counter:03d}",
            "investor_name": inv["investor_name"],
            "principal_amount": inv["principal_amount"],
            "investment_date": inv["investment_date"],
            "year": inv["year"],
            "status": inv["status"],
            "bf_amount": inv["bf_amount"],
        })

    with open(inv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"\nInvestments CSV written to: {inv_path}")
    print(f"  {len(rows)} investment records from {len(set(r['year'] for r in rows))} years")

    # Summary by year
    by_year = defaultdict(lambda: {"count": 0, "total": 0})
    for r in rows:
        by_year[r["year"]]["count"] += 1
        by_year[r["year"]]["total"] += r["principal_amount"]
    for yr in sorted(by_year):
        info = by_year[yr]
        print(f"    {yr}: {info['count']} investors, K {info['total']:,.2f} total")


def generate_interest_expenses_csv(wb):
    """
    Extract INTEREST EXPENSE tables from all sheets of jutem_fund.xlsx.

    Each sheet has an 'INTEREST EXPENSE' header followed by:
        NAME | DATE EXPECTED | BORROWED | INTEREST RATE | PAID | INTEREST EXPENSE | STATUS

    Formulas are computed in Python:
        paid = borrowed * (1 + interest_rate)
        interest_expense = borrowed * interest_rate

    Output: interest_expenses_upload.csv
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    ie_path = os.path.join(OUTPUT_DIR, "interest_expenses_upload.csv")

    fieldnames = [
        "investor_name", "date_expected", "borrowed_amount", "interest_rate",
        "paid_amount", "interest_expense", "reference_number", "sheet",
    ]

    all_rows = []
    ref_counter = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        year, month = parse_sheet_date(sheet_name)
        if year is None:
            continue

        # Find the "INTEREST EXPENSE" header row
        ie_start = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=1, values_only=False):
            cell_val = row[0].value
            if cell_val and str(cell_val).strip().upper() == "INTEREST EXPENSE":
                ie_start = row[0].row
                break

        if ie_start is None:
            continue

        # Skip the column header row (NAME | DATE EXPECTED | ...)
        data_start = ie_start + 2

        for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row, max_col=7, values_only=False):
            cells = [c.value for c in row]
            name = cells[0]

            # Stop at TOTAL or empty name
            if not name:
                break
            name_str = str(name).strip()
            if name_str.upper() in ("TOTAL", ""):
                break

            date_val = cells[1]
            borrowed = cells[2]
            rate = cells[3]

            # Skip rows with missing key values
            if borrowed is None or rate is None:
                continue

            try:
                borrowed = float(borrowed)
                rate = float(rate)
            except (ValueError, TypeError):
                continue

            if borrowed <= 0:
                continue

            # Compute: FV(rate, 1, 0, -borrowed) = borrowed * (1 + rate)
            paid = round(borrowed * (1 + rate), 2)
            interest_exp = round(borrowed * rate, 2)

            # Parse date
            txn_date = None
            if isinstance(date_val, datetime):
                txn_date = date_val.strftime("%Y-%m-%d")
            elif date_val:
                try:
                    txn_date = datetime.strptime(str(date_val).strip(), "%Y-%m-%d").strftime("%Y-%m-%d")
                except ValueError:
                    pass
            if not txn_date:
                txn_date = f"{year}-{month:02d}-01"

            ref_counter += 1
            all_rows.append({
                "investor_name": name_str,
                "date_expected": txn_date,
                "borrowed_amount": borrowed,
                "interest_rate": rate,
                "paid_amount": paid,
                "interest_expense": interest_exp,
                "reference_number": f"INTEXP-{ref_counter:05d}",
                "sheet": sheet_name,
            })

    with open(ie_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_rows)

    print(f"\nInterest Expenses CSV written to: {ie_path}")
    print(f"  {len(all_rows)} interest expense rows from {len(wb.sheetnames)} sheets")

    # Summary by investor
    by_investor = defaultdict(lambda: {"count": 0, "total": 0})
    for r in all_rows:
        by_investor[r["investor_name"]]["count"] += 1
        by_investor[r["investor_name"]]["total"] += r["interest_expense"]
    print(f"  Top interest expense payers:")
    for name, info in sorted(by_investor.items(), key=lambda x: -x[1]["total"])[:10]:
        print(f"    {name}: {info['count']} payments, K {info['total']:,.2f} total interest")


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Jutem Fund ETL")
    parser.add_argument(
        "--step",
        choices=["review", "borrowers", "loans", "payments", "expenses",
                 "investors", "investments", "interest_expenses", "all"],
        default="all",
        help="Which step to run (default: all)"
    )
    args = parser.parse_args()

    wb = load_workbook()
    all_loans = extract_all_loans(wb)
    registry = build_borrower_registry(all_loans)

    if args.step == "review":
        generate_review_csv(registry)
        return

    if args.step in ("borrowers", "all"):
        output_path, nrc_map = generate_borrowers_csv(registry)

    if args.step in ("loans", "all"):
        # Load NRC map from file if only running loans step
        if args.step == "loans":
            mapping_path = os.path.join(OUTPUT_DIR, "name_to_nrc.csv")
            if not os.path.exists(mapping_path):
                print(f"ERROR: {mapping_path} not found. Run --step borrowers first.")
                sys.exit(1)
            nrc_map = {}
            with open(mapping_path, "r") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    nrc_map[row["canonical_name"]] = row["nrc"]
            print(f"Loaded {len(nrc_map)} NRC mappings")

        generate_loans_csv(all_loans, nrc_map)

    if args.step in ("payments", "all"):
        generate_payments_csv()

    if args.step in ("expenses", "all"):
        generate_expenses_csv(wb)

    if args.step in ("investors", "all"):
        generate_investors_csv()

    if args.step in ("investments", "all"):
        generate_investments_csv()

    if args.step in ("interest_expenses", "all"):
        generate_interest_expenses_csv(wb)

    print("\n✓ ETL complete!")
    print(f"\nNext steps:")
    print(f"  1. Upload borrowers:          etl/output/borrowers_upload.csv")
    print(f"  2. Upload loans:              etl/output/loans_upload.csv")
    print(f"  3. Upload amortization:       etl/output/amortization_upload.csv")
    print(f"  4. Upload payments:           etl/output/payments_upload.csv")
    print(f"  5. Upload expenses:           etl/output/expenses_upload.csv")
    print(f"  6. Upload investors:          etl/output/investors_upload.csv")
    print(f"  7. Upload investments:        etl/output/investments_upload.csv")
    print(f"  8. Upload interest expenses:  etl/output/interest_expenses_upload.csv")


if __name__ == "__main__":
    main()
