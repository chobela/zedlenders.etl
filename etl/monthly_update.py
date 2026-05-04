#!/usr/bin/env python3
"""
Jutem Fund Monthly Delta Update ETL
====================================
Generic version of march_2026_update.py. Processes ANY month's sheet from
jutem_fund.xlsx (or a standalone month workbook) as a delta update.

The sheet name (e.g. "APRIL 2026") drives year, month, date ranges, and
reference-number prefixes (e.g. "PMT-APR26-...").

Usage:
    python3 etl/monthly_update.py --sheet "APRIL 2026"
    python3 etl/monthly_update.py --sheet "APRIL 2026" --execute
    python3 etl/monthly_update.py --sheet "APRIL 2026" --step loans
    python3 etl/monthly_update.py --workbook /path/to/upload.xlsx --sheet "APRIL 2026"
"""

import os
import sys
import re
import csv
import json
import argparse
import time
import calendar
import requests
from datetime import datetime
from collections import defaultdict

# ─── Static configuration ────────────────────────────────────────────────────

DIRECTUS_URL = "https://zedlenders.pickmesms.com"
COMPANY_ID = 22
ACCOUNT_ID = 25  # Jutem Fund account
DEFAULT_WORKBOOK = os.path.join(os.path.dirname(__file__), "..", "jutem_fund.xlsx")
NRC_MAPPING_PATH = os.path.join(os.path.dirname(__file__), "output", "name_to_nrc.csv")

# ─── Per-run configuration (set in main() from CLI args) ─────────────────────
# Module-level so existing helper functions can read them without rewiring.

EXCEL_FILE = None        # Path to .xlsx workbook
SHEET_NAME = None        # Excel sheet to process, e.g. "APRIL 2026"
YEAR = None              # int, e.g. 2026
MONTH = None             # int 1-12
MONTH_NAME = None        # canonical "APRIL 2026" (uppercase)
MONTH_ABBREV = None      # e.g. "APR26" — used for reference-number prefixes
OUTPUT_DIR = None        # etl/output/<slug>


def configure_run(sheet_name, workbook_path):
    """Set all per-run module globals. Call once at the start of main()."""
    global EXCEL_FILE, SHEET_NAME, YEAR, MONTH, MONTH_NAME, MONTH_ABBREV, OUTPUT_DIR

    EXCEL_FILE = workbook_path

    parts = sheet_name.strip().split()
    if len(parts) != 2 or not parts[1].isdigit():
        raise ValueError(
            f"--sheet must be 'MONTH YEAR' (e.g. 'APRIL 2026'), got: {sheet_name!r}"
        )
    try:
        dt = datetime.strptime(sheet_name.strip().title(), "%B %Y")
    except ValueError as e:
        raise ValueError(f"Could not parse month name in {sheet_name!r}: {e}")

    YEAR = dt.year
    MONTH = dt.month
    MONTH_NAME = sheet_name.strip().upper()
    SHEET_NAME = sheet_name.strip()
    MONTH_ABBREV = f"{dt.strftime('%b').upper()}{str(YEAR)[-2:]}"  # e.g. APR26

    slug = f"{YEAR}_{MONTH:02d}_{dt.strftime('%b').lower()}"
    OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "output", slug)


def month_start_iso():
    return f"{YEAR}-{MONTH:02d}-01"


def month_end_iso():
    last_day = calendar.monthrange(YEAR, MONTH)[1]
    return f"{YEAR}-{MONTH:02d}-{last_day:02d}"

# ─── Name normalization (from jutem_etl.py) ──────────────────────────────────

NAME_ALIASES = {
    "Ba Chrsitabel": "Ba Christabel",
    "Ba Chrisitabel": "Ba Christabel",
    "Amingtone": "Amington",
    "Amingtone Musonda": "Amington Musonda",
    "KONDWA": "Kondwa",
    "A.milimo": "Milimo",
    "A.Milimo": "Milimo",
    "Mwaka Milimo": "Mwaka Milimo",
    "A.B.M (Kwezekani)": "ABM Kwezekani",
    "Jeffrey Cimankata": "Jeffrey Chimankata",
    "Money Acumen (Lyapa and Theresa)": "Money Acumen Lyapa and Theresa",
}

SKIP_EXACT = {
    "TOTAL", "NAME", "DATE", "COMMISSION", "BORROWED",
    "INTEREST INCOME", "INTEREST EXPENSE", "NET INTEREST INCOME",
    "NET OPERATING INCOME", "OTHER EXPENSES", "TOTAL EXPENSES",
    "NET INTEREST INCOME AND TRACKING FEE", "TRACKING FEES",
}


def clean_name(raw_name):
    if not raw_name or not isinstance(raw_name, str):
        return None
    name = raw_name.strip()
    if name in NAME_ALIASES:
        name = NAME_ALIASES[name]
    name = re.sub(r"\s+", " ", name).strip()
    return name


def split_first_last(name):
    parts = name.strip().split()
    if len(parts) == 0:
        return ("Unknown", "N/A")
    elif len(parts) == 1:
        return (parts[0], "N/A")
    else:
        return (parts[0], " ".join(parts[1:]))


def parse_day(text):
    if not text:
        return None
    text = str(text).strip()
    match = re.search(r"(\d{1,2})", text)
    if match:
        day = int(match.group(1))
        if 1 <= day <= 31:
            max_day = calendar.monthrange(YEAR, MONTH)[1]
            return min(day, max_day)
    return None


def make_date(day):
    if day:
        max_day = calendar.monthrange(YEAR, MONTH)[1]
        day = min(day, max_day)
        return f"{YEAR}-{MONTH:02d}-{day:02d}"
    return f"{YEAR}-{MONTH:02d}-01"


# ─── Token ───────────────────────────────────────────────────────────────────

def get_token():
    token = os.environ.get("DIRECTUS_TOKEN")
    if not token:
        token = input("Enter Directus admin token: ").strip()
    return token


# ─── Load Excel ──────────────────────────────────────────────────────────────

def load_excel():
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required. pip install openpyxl")
        sys.exit(1)

    path = os.path.abspath(EXCEL_FILE)
    print(f"Loading: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)

    # Match SHEET_NAME case-insensitively against available sheets.
    target = SHEET_NAME.strip().upper()
    match = next((s for s in wb.sheetnames if s.strip().upper() == target), None)
    if match is None:
        raise ValueError(
            f"Sheet {SHEET_NAME!r} not found in workbook. "
            f"Available sheets: {wb.sheetnames}"
        )
    ws = wb[match]
    print(f"Sheet: {ws.title}, {ws.max_row} rows x {ws.max_column} cols")
    return ws


def parse_all_sections(ws):
    """Parse the Excel into: loans, interest_expenses, tracking_fees, other_expenses."""
    loans = []
    interest_expenses = []
    tracking_fees = []
    other_expenses = []

    current_section = None
    skip_next_header = False

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        first_cell = str(row[0]).strip() if row[0] else ""
        upper = first_cell.upper()

        # Section headers
        if upper == "NAME" and current_section is None:
            current_section = "loans"
            continue
        elif upper == "INTEREST EXPENSE":
            current_section = "interest_expense"
            skip_next_header = True
            continue
        elif upper == "TRACKING FEES":
            current_section = "tracking"
            skip_next_header = True
            continue
        elif upper == "COMMISSION":
            current_section = "commission"
            skip_next_header = True
            continue
        elif upper == "OTHER EXPENSES":
            current_section = "expenses"
            skip_next_header = True
            continue
        elif upper in ("NET INTEREST INCOME", "NET INTEREST INCOME AND TRACKING FEE",
                        "NET OPERATING INCOME", "PROFIT BEFORE TAX"):
            current_section = "summary"
            continue

        if skip_next_header:
            skip_next_header = False
            continue

        if upper in SKIP_EXACT or upper.startswith("INTEREST INCOME"):
            if current_section == "loans":
                current_section = None
            continue

        if upper == "TOTAL":
            current_section = None
            continue

        if not first_cell or all(v is None for v in row[:3]):
            continue

        if current_section == "loans":
            name = first_cell
            amount = row[1] if len(row) > 1 else None
            interest = row[2] if len(row) > 2 else None
            expected = row[3] if len(row) > 3 else None
            date_text = row[4] if len(row) > 4 else None
            profit = row[5] if len(row) > 5 else None
            status = row[6] if len(row) > 6 else None
            commission = row[7] if len(row) > 7 else None

            if not amount or not isinstance(amount, (int, float)):
                continue

            canonical = clean_name(name)
            if not canonical:
                continue

            interest_val = float(interest) if interest and isinstance(interest, (int, float)) else 0
            interest_pct = round(interest_val * 100, 2)
            expected_val = float(expected) if expected and isinstance(expected, (int, float)) else round(float(amount) * (1 + interest_val), 2)
            profit_val = float(profit) if profit and isinstance(profit, (int, float)) else round(expected_val - float(amount), 2)

            status_str = str(status).strip().lower() if status else "pending"
            day = parse_day(date_text)

            loans.append({
                "raw_name": name,
                "canonical_name": canonical,
                "amount": round(float(amount), 2),
                "interest_pct": interest_pct,
                "interest_decimal": interest_val,
                "expected_amount": round(expected_val, 2),
                "profit": round(profit_val, 2),
                "due_date": make_date(day),
                "day": day,
                "status": status_str,
                "commission": float(commission) if commission and isinstance(commission, (int, float)) else 0,
            })

        elif current_section == "interest_expense":
            name = first_cell
            date_val = row[1] if len(row) > 1 else None
            borrowed = row[2] if len(row) > 2 else None
            rate = row[3] if len(row) > 3 else None

            if not borrowed or not isinstance(borrowed, (int, float)):
                continue

            rate_val = float(rate) if rate and isinstance(rate, (int, float)) else 0
            borrowed_val = float(borrowed)
            paid_val = round(borrowed_val * (1 + rate_val), 2)
            ie_val = round(borrowed_val * rate_val, 2)

            txn_date = None
            if isinstance(date_val, datetime):
                txn_date = date_val.strftime("%Y-%m-%d")
            if not txn_date:
                txn_date = f"{YEAR}-{MONTH:02d}-01"

            interest_expenses.append({
                "investor_name": name.strip(),
                "date_expected": txn_date,
                "borrowed_amount": round(borrowed_val, 2),
                "interest_rate": rate_val,
                "paid_amount": paid_val,
                "interest_expense": ie_val,
            })

        elif current_section == "tracking":
            name = first_cell
            date_val = row[1] if len(row) > 1 else None
            fee = row[2] if len(row) > 2 else None

            if not fee or not isinstance(fee, (int, float)) or float(fee) <= 0:
                continue

            txn_date = None
            if isinstance(date_val, datetime):
                txn_date = date_val.strftime("%Y-%m-%d")
            if not txn_date:
                txn_date = f"{YEAR}-{MONTH:02d}-01"

            tracking_fees.append({
                "name": name.strip(),
                "date": txn_date,
                "fee": round(float(fee), 2),
            })

        elif current_section == "expenses":
            date_val = row[0]
            description = row[1] if len(row) > 1 else None
            amount = row[2] if len(row) > 2 else None

            if not description or not amount:
                continue

            try:
                amount_val = float(amount)
            except (ValueError, TypeError):
                continue

            if amount_val <= 0:
                continue

            txn_date = None
            if isinstance(date_val, datetime):
                txn_date = date_val.strftime("%Y-%m-%d")
            elif date_val and str(date_val).strip() not in ("-", ""):
                try:
                    txn_date = datetime.strptime(str(date_val).strip(), "%Y-%m-%d").strftime("%Y-%m-%d")
                except ValueError:
                    pass
            if not txn_date:
                txn_date = f"{YEAR}-{MONTH:02d}-01"

            other_expenses.append({
                "description": str(description).strip(),
                "amount": round(amount_val, 2),
                "date": txn_date,
            })

    return loans, interest_expenses, tracking_fees, other_expenses


# ─── Directus API helpers ────────────────────────────────────────────────────

def api_get(token, path, params=None):
    headers = {"Authorization": f"Bearer {token}"}
    res = requests.get(f"{DIRECTUS_URL}{path}", headers=headers, params=params or {})
    res.raise_for_status()
    return res.json().get("data", [])


def api_post(token, path, data):
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    res = requests.post(f"{DIRECTUS_URL}{path}", headers=headers, json=data)
    res.raise_for_status()
    return res.json().get("data", {})


def api_patch(token, path, data):
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    res = requests.patch(f"{DIRECTUS_URL}{path}", headers=headers, json=data)
    res.raise_for_status()
    return res.json().get("data", {})


# ─── Fetch existing data ────────────────────────────────────────────────────

def fetch_borrowers(token):
    """Fetch all company 22 borrowers, indexed by lowercase full name and by NRC."""
    print("Fetching existing borrowers...")
    borrowers = api_get(token, "/users", {
        "filter[role][name][_eq]": "Borrower",
        "filter[company][_eq]": COMPANY_ID,
        "fields": "id,first_name,last_name,nrc,phone",
        "limit": -1,
    })
    by_name = {}
    by_nrc = {}
    for b in borrowers:
        full = f"{b.get('first_name', '')} {b.get('last_name', '')}".strip().lower()
        by_name[full] = b
        if b.get("nrc"):
            by_nrc[b["nrc"]] = b
    print(f"  {len(borrowers)} borrowers loaded")
    return by_name, by_nrc


def fetch_month_amortization(token):
    """Fetch ALL amortization entries for the current month (no company field on this table)."""
    print(f"Fetching existing {MONTH_NAME} amortization entries...")
    entries = api_get(token, "/items/amortization", {
        "filter[due_date][_gte]": month_start_iso(),
        "filter[due_date][_lte]": month_end_iso(),
        "fields": "*",
        "limit": -1,
    })
    print(f"  {len(entries)} existing amortization entries for {MONTH_NAME}")
    return entries


def fetch_loans_for_company(token):
    """Fetch all company 22 loans."""
    print("Fetching existing loans...")
    loans = api_get(token, "/items/loans", {
        "filter[company][_eq]": COMPANY_ID,
        "fields": "id,borrower.id,borrower.first_name,borrower.last_name,borrower.nrc,amount,custom_interest_rate,loan_status,approval_date",
        "limit": -1,
    })
    print(f"  {len(loans)} loans loaded")
    return loans


def fetch_loan_product(token):
    products = api_get(token, "/items/loan_products", {
        "filter[company][_eq]": COMPANY_ID,
        "filter[is_active][_eq]": "true",
        "fields": "id,loan_name",
        "limit": -1,
    })
    for p in products:
        if "micro" in p["loan_name"].lower() or "business" in p["loan_name"].lower():
            return p["id"]
    return products[0]["id"] if products else None


def fetch_branch(token):
    branches = api_get(token, "/items/branches", {
        "filter[company][_eq]": COMPANY_ID,
        "fields": "id,branch_name",
        "limit": -1,
    })
    for b in branches:
        if "lusaka" in b["branch_name"].lower():
            return b["id"]
    return branches[0]["id"] if branches else None


def fetch_borrower_role_id(token):
    roles = api_get(token, "/roles", {"filter[name][_eq]": "Borrower", "fields": "id,name"})
    return roles[0]["id"] if roles else None


def load_nrc_mapping():
    mapping = {}
    if os.path.exists(NRC_MAPPING_PATH):
        with open(NRC_MAPPING_PATH, "r") as f:
            reader = csv.DictReader(f)
            for row in reader:
                mapping[row["canonical_name"].lower()] = row["nrc"]
    print(f"  Loaded {len(mapping)} NRC mappings from initial ETL")
    return mapping


# ─── Borrower matching ──────────────────────────────────────────────────────

def find_borrower(canonical, by_name, by_nrc, nrc_map):
    """Find an existing borrower by NRC or name. Returns borrower_id or None."""
    canonical_lower = canonical.lower()

    # Try NRC first
    nrc = nrc_map.get(canonical_lower) or nrc_map.get(canonical)
    if nrc and nrc in by_nrc:
        return by_nrc[nrc]["id"]

    # Exact name match
    if canonical_lower in by_name:
        return by_name[canonical_lower]["id"]

    # Partial match (the name is contained in or contains an existing name)
    for name_key, b in by_name.items():
        if canonical_lower in name_key or name_key in canonical_lower:
            return b["id"]

    return None


# ─── Step 1: Process Loans (the big one) ────────────────────────────────────

def process_loans(loans, token, execute=False):
    """
    Delta update for March 2026 loans:

    A) For 83 EXISTING amortization entries:
       - Fill in expected_amount and profit (were None from original upload)
       - Update status pending → paid for 29 entries
       - Create payment records for newly-paid entries
       - Update parent loan status if all installments paid

    B) For 17 NEW loans:
       - Find or create borrower
       - Create loan record
       - Create amortization entry
       - Create payment if status = paid
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    by_name, by_nrc = fetch_borrowers(token)
    nrc_map = load_nrc_mapping()
    existing_loans = fetch_loans_for_company(token)
    existing_amort = fetch_month_amortization(token)
    product_id = fetch_loan_product(token)
    branch_id = fetch_branch(token)
    role_id = fetch_borrower_role_id(token)

    print(f"\n  Product ID: {product_id}, Branch: {branch_id}, Borrower Role: {role_id}")

    # Build amortization index: (borrower_nrc, amount_rounded) → amort entry
    # The amortization table has borrower_nrc directly, no need for relational joins
    amort_by_nrc_amount = {}
    for ea in existing_amort:
        nrc = ea.get("borrower_nrc", "")
        amount_key = round(float(ea.get("amount_due", 0)), 0)
        amort_by_nrc_amount[(nrc, amount_key)] = ea

    # Also build a reverse NRC → name map from existing borrowers for logging
    nrc_to_name = {}
    for name_key, b in by_name.items():
        if b.get("nrc"):
            nrc_to_name[b["nrc"]] = name_key

    # Build loans index by borrower_id
    loans_by_borrower = defaultdict(list)
    for el in existing_loans:
        bid = el.get("borrower")
        if isinstance(bid, dict):
            bid = bid["id"]
        if bid:
            loans_by_borrower[bid].append(el)

    stats = {
        "amort_updated": 0,        # expected_amount/profit filled in
        "status_to_paid": 0,       # pending → paid
        "payments_created": 0,
        "loans_status_updated": 0, # loan status active → settled
        "new_borrowers": 0,
        "new_loans": 0,
        "new_amortization": 0,
        "matched_existing": 0,
        "errors": [],
    }

    # CSV audit logs
    log_files = {}
    for name, headers in [
        ("amort_updates", ["amort_id", "borrower", "amount_due", "field", "old_value", "new_value"]),
        ("new_payments", ["loan_id", "amort_id", "borrower", "amount", "date", "reference"]),
        ("new_loans", ["borrower_name", "borrower_id", "amount", "interest", "expected", "due_date", "status"]),
        ("new_borrowers", ["name", "first_name", "last_name"]),
        ("loan_status_updates", ["loan_id", "borrower", "old_status", "new_status"]),
    ]:
        f = open(os.path.join(OUTPUT_DIR, f"{name}.csv"), "w", newline="")
        w = csv.writer(f)
        w.writerow(headers)
        log_files[name] = (f, w)

    pmt_counter = 0

    for idx, loan_row in enumerate(loans):
        canonical = loan_row["canonical_name"]
        canonical_lower = canonical.lower()
        amount_key = round(loan_row["amount"], 0)

        # ── A) Try to match to an existing amortization entry via NRC ──
        # First find this borrower's NRC
        borrower_nrc = nrc_map.get(canonical_lower) or nrc_map.get(canonical)
        matched_amort = amort_by_nrc_amount.get((borrower_nrc, amount_key)) if borrower_nrc else None

        if matched_amort:
            amort_id = matched_amort["id"]
            loan_ref = matched_amort.get("loan_reference")
            loan_id = loan_ref["id"] if isinstance(loan_ref, dict) else loan_ref
            stats["matched_existing"] += 1

            # A1) Update expected_amount and profit if currently empty
            old_expected = matched_amort.get("expected_amount")
            old_profit = matched_amort.get("profit")
            updates = {}

            if not old_expected or float(old_expected or 0) == 0:
                updates["expected_amount"] = loan_row["expected_amount"]
                log_files["amort_updates"][1].writerow([
                    amort_id, canonical, loan_row["amount"],
                    "expected_amount", old_expected, loan_row["expected_amount"]
                ])

            if not old_profit or float(old_profit or 0) == 0:
                updates["profit"] = loan_row["profit"]
                log_files["amort_updates"][1].writerow([
                    amort_id, canonical, loan_row["amount"],
                    "profit", old_profit, loan_row["profit"]
                ])

            # A2) Update status if changed to paid
            old_status = (matched_amort.get("status") or "pending").lower()
            new_status = loan_row["status"]

            if old_status != new_status and new_status in ("paid", "cleared"):
                updates["status"] = new_status
                log_files["amort_updates"][1].writerow([
                    amort_id, canonical, loan_row["amount"],
                    "status", old_status, new_status
                ])
                stats["status_to_paid"] += 1

            if updates:
                stats["amort_updated"] += 1
                if execute:
                    try:
                        api_patch(token, f"/items/amortization/{amort_id}", updates)
                        fields = ", ".join(f"{k}={v}" for k, v in updates.items())
                        print(f"  ~ Updated amort {amort_id} ({canonical}): {fields}")
                        time.sleep(0.1)
                    except Exception as e:
                        stats["errors"].append(f"Update amort {amort_id}: {e}")
                        print(f"  ! Error updating amort {amort_id}: {e}")
                else:
                    fields = ", ".join(f"{k}={v}" for k, v in updates.items())
                    print(f"  [DRY] Would update amort {amort_id} ({canonical}): {fields}")

            # A3) Create payment for newly-paid entries
            if old_status != new_status and new_status in ("paid", "cleared"):
                pmt_counter += 1
                ref = f"PMT-{MONTH_ABBREV}-{pmt_counter:05d}"
                log_files["new_payments"][1].writerow([
                    loan_id, amort_id, canonical,
                    loan_row["expected_amount"], loan_row["due_date"], ref
                ])
                stats["payments_created"] += 1

                if execute:
                    try:
                        api_post(token, "/items/transactions", {
                            "loan": loan_id,
                            "amount": loan_row["expected_amount"],
                            "transaction_date": loan_row["due_date"],
                            "transaction_type": "Loan Repayment",
                            "payment_method": "cash",
                            "reference_number": ref,
                            "transaction_status": "Completed",
                            "company": COMPANY_ID,
                            "amortization_installment": amort_id,
                        })
                        print(f"  + Payment: {canonical} K{loan_row['expected_amount']:,.2f}")
                        time.sleep(0.1)
                    except Exception as e:
                        stats["errors"].append(f"Payment {canonical}: {e}")
                        print(f"  ! Error creating payment for {canonical}: {e}")
                else:
                    print(f"  [DRY] Would create payment: {canonical} K{loan_row['expected_amount']:,.2f}")

                # A4) Update parent loan's paid_amount and possibly status
                if execute and loan_id:
                    try:
                        # Fetch current loan to update paid_amount
                        loan_data = api_get(token, f"/items/loans/{loan_id}", {
                            "fields": "id,paid_amount,loan_status,amount"
                        })
                        if isinstance(loan_data, list):
                            loan_data = loan_data[0] if loan_data else {}
                        current_paid = float(loan_data.get("paid_amount", 0) or 0)
                        new_paid = current_paid + loan_row["expected_amount"]
                        loan_update = {"paid_amount": round(new_paid, 2)}

                        # Check if loan is now fully paid
                        loan_amount = float(loan_data.get("amount", 0) or 0)
                        if new_paid >= loan_amount and loan_data.get("loan_status") == 5:
                            loan_update["loan_status"] = 6  # Settled
                            stats["loans_status_updated"] += 1
                            log_files["loan_status_updates"][1].writerow([
                                loan_id, canonical, 5, 6
                            ])

                        api_patch(token, f"/items/loans/{loan_id}", loan_update)
                        time.sleep(0.1)
                    except Exception as e:
                        stats["errors"].append(f"Update loan {loan_id}: {e}")

            continue  # Done with this matched entry

        # ── B) NEW LOAN — not in existing amortization ──
        borrower_id = find_borrower(canonical, by_name, by_nrc, nrc_map)

        if not borrower_id:
            # Create new borrower
            first, last = split_first_last(canonical)
            nrc_num = 1001000 + idx
            new_nrc = f"{nrc_num:07d}/00/1"
            phone = f"26097700{1000+idx}"
            email = f"{nrc_num}@kwachaplus.com"

            log_files["new_borrowers"][1].writerow([canonical, first, last])
            stats["new_borrowers"] += 1

            if execute and role_id:
                try:
                    new_user = api_post(token, "/users", {
                        "first_name": first,
                        "last_name": last,
                        "email": email,
                        "password": "JutemBorrower2026!",
                        "nrc": new_nrc,
                        "phone": phone,
                        "role": role_id,
                        "company": COMPANY_ID,
                        "province": "Lusaka",
                        "address": "Lusaka",
                        "employment_type": "self_employed",
                    })
                    borrower_id = new_user["id"]
                    by_name[canonical_lower] = new_user
                    by_nrc[new_nrc] = new_user
                    print(f"  + Created borrower: {canonical} (ID: {borrower_id})")
                    time.sleep(0.2)
                except Exception as e:
                    stats["errors"].append(f"Create borrower {canonical}: {e}")
                    print(f"  ! Error creating borrower {canonical}: {e}")
                    continue
            else:
                print(f"  [DRY] Would create borrower: {canonical}")
                if not execute:
                    continue

        # Create loan
        loan_status = 6 if loan_row["status"] in ("paid", "cleared") else 5

        log_files["new_loans"][1].writerow([
            canonical, borrower_id, loan_row["amount"], loan_row["interest_pct"],
            loan_row["expected_amount"], loan_row["due_date"], loan_row["status"]
        ])
        stats["new_loans"] += 1

        new_loan_id = None
        new_amort_id = None

        if execute and product_id and branch_id:
            try:
                new_loan = api_post(token, "/items/loans", {
                    "borrower": borrower_id,
                    "loan_product": product_id,
                    "amount": loan_row["amount"],
                    "custom_interest_rate": loan_row["interest_pct"],
                    "loan_status": loan_status,
                    "application_date": f"{YEAR}-{MONTH:02d}-01",
                    "approval_date": f"{YEAR}-{MONTH:02d}-01",
                    "loan_purpose": "Business",
                    "branch": branch_id,
                    "company": COMPANY_ID,
                    "loan_term": 1,
                    "paid_amount": loan_row["expected_amount"] if loan_status == 6 else 0,
                })
                new_loan_id = new_loan["id"]
                print(f"  + New loan: {canonical} K{loan_row['amount']:,.2f} (ID: {new_loan_id})")
                time.sleep(0.2)
            except Exception as e:
                stats["errors"].append(f"Create loan {canonical}: {e}")
                print(f"  ! Error creating loan for {canonical}: {e}")
                continue

            # Create amortization
            # Find borrower NRC for this new loan
            new_borrower_nrc = ""
            for nrc_val, b in by_nrc.items():
                if b["id"] == borrower_id:
                    new_borrower_nrc = nrc_val
                    break

            try:
                new_amort = api_post(token, "/items/amortization", {
                    "loan_reference": new_loan_id,
                    "borrower_nrc": new_borrower_nrc,
                    "month": MONTH_NAME,
                    "due_date": loan_row["due_date"],
                    "amount_due": loan_row["amount"],
                    "interest_rate": loan_row["interest_pct"],
                    "expected_amount": loan_row["expected_amount"],
                    "profit": loan_row["profit"],
                    "status": loan_row["status"],
                    "sheet": MONTH_NAME,
                })
                new_amort_id = new_amort.get("id")
                stats["new_amortization"] += 1
                print(f"  + Amortization: {canonical} ({loan_row['status']})")
                time.sleep(0.1)
            except Exception as e:
                stats["errors"].append(f"Create amort {canonical}: {e}")

            # Create payment if paid
            if loan_status == 6 and new_loan_id:
                pmt_counter += 1
                ref = f"PMT-{MONTH_ABBREV}-{pmt_counter:05d}"
                try:
                    api_post(token, "/items/transactions", {
                        "loan": new_loan_id,
                        "amount": loan_row["expected_amount"],
                        "transaction_date": loan_row["due_date"],
                        "transaction_type": "Loan Repayment",
                        "payment_method": "cash",
                        "reference_number": ref,
                        "transaction_status": "Completed",
                        "company": COMPANY_ID,
                        "amortization_installment": new_amort_id,
                    })
                    stats["payments_created"] += 1
                    print(f"  + Payment: {canonical} K{loan_row['expected_amount']:,.2f}")
                    time.sleep(0.1)
                except Exception as e:
                    stats["errors"].append(f"Payment {canonical}: {e}")
        else:
            print(f"  [DRY] Would create loan + amort: {canonical} K{loan_row['amount']:,.2f} ({loan_row['status']})")
            stats["new_amortization"] += 1
            if loan_status == 6:
                stats["payments_created"] += 1

    # Close log files
    for name, (f, w) in log_files.items():
        f.close()

    return stats


# ─── Step 2: Process Interest Expenses ──────────────────────────────────────

def process_interest_expenses(interest_expenses, token, execute=False):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    stats = {"created": 0, "skipped": 0, "errors": []}

    # Interest expenses are stored in `transactions` with transaction_type="interest_expense"
    existing = api_get(token, "/items/transactions", {
        "filter[company][_eq]": COMPANY_ID,
        "filter[transaction_type][_eq]": "interest_expense",
        "filter[transaction_date][_gte]": month_start_iso(),
        "filter[transaction_date][_lte]": month_end_iso(),
        "fields": "id,reference_number,notes,amount",
        "limit": -1,
    })
    existing_refs = {e.get("reference_number", "") for e in existing}
    print(f"  {len(existing)} existing {MONTH_NAME} interest expense transactions")

    # Investor lookup for investment linking
    investors = api_get(token, "/items/jutem_investors", {
        "filter[company][_eq]": COMPANY_ID,
        "fields": "id,name",
        "limit": -1,
    })
    investor_map = {inv["name"].upper().strip(): inv["id"] for inv in investors}

    # Load investment mapping for linking
    inv_mapping_path = os.path.join(os.path.dirname(__file__), "output", "investment_ref_to_id.csv")
    inv_mapping = {}
    if os.path.exists(inv_mapping_path):
        with open(inv_mapping_path, "r") as mf:
            reader = csv.DictReader(mf)
            for row in reader:
                key = (row["investor_name"].upper().strip(), int(row["year"]))
                inv_mapping[key] = int(row["id"])
        print(f"  Loaded {len(inv_mapping)} investment mappings")

    ie_csv = os.path.join(OUTPUT_DIR, "interest_expenses.csv")
    with open(ie_csv, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["investor_name", "date", "borrowed", "rate", "paid", "interest_expense", "action"])

        # Fetch account balance for running balance
        account_balance = 0
        if execute:
            try:
                acct = api_get(token, f"/items/accounts/{ACCOUNT_ID}", {"fields": "id,balance"})
                if isinstance(acct, list):
                    acct = acct[0] if acct else {}
                account_balance = float(acct.get("balance", 0) or 0)
                print(f"  Account {ACCOUNT_ID} balance: K {account_balance:,.2f}")
            except Exception:
                pass
        running_balance = account_balance

        for idx, ie in enumerate(interest_expenses):
            ref = f"INTEXP-{MONTH_ABBREV}-{idx+1:04d}"

            if ref in existing_refs:
                writer.writerow([ie["investor_name"], ie["date_expected"], ie["borrowed_amount"],
                                ie["interest_rate"], ie["paid_amount"], ie["interest_expense"], "SKIP"])
                stats["skipped"] += 1
                continue

            writer.writerow([ie["investor_name"], ie["date_expected"], ie["borrowed_amount"],
                            ie["interest_rate"], ie["paid_amount"], ie["interest_expense"], "CREATE"])
            stats["created"] += 1

            # Look up investment_id for linking
            inv_key = (ie["investor_name"].upper().strip(), YEAR)
            investment_id = inv_mapping.get(inv_key)

            if execute:
                running_balance -= ie["paid_amount"]
                try:
                    payload = {
                        "amount": ie["interest_expense"],
                        "transfer_fees": 0,
                        "transaction_date": ie["date_expected"],
                        "payment_method": "cash",
                        "transaction_type": "interest_expense",
                        "reference_number": ref,
                        "notes": (
                            f"Interest payment to {ie['investor_name']} | "
                            f"Borrowed: K{ie['borrowed_amount']:,.2f} | "
                            f"Rate: {ie['interest_rate']*100:.1f}% | "
                            f"Interest: K{ie['interest_expense']:,.2f}"
                        ),
                        "transaction_status": "Completed",
                        "is_loan_transaction": False,
                        "is_debt_repayment": True,
                        "is_gps_fee_payment": False,
                        "new_amount": round(running_balance, 2),
                        "company": COMPANY_ID,
                        "account": ACCOUNT_ID,
                    }
                    if investment_id:
                        payload["investment_reference"] = investment_id
                    api_post(token, "/items/transactions", payload)
                    print(f"  + IE: {ie['investor_name']} K{ie['interest_expense']:,.2f}")
                    time.sleep(0.1)
                except Exception as e:
                    stats["errors"].append(f"IE {ie['investor_name']}: {e}")
                    print(f"  ! Error: {ie['investor_name']}: {e}")
            else:
                print(f"  [DRY] Would create IE: {ie['investor_name']} K{ie['interest_expense']:,.2f}")

        # Update account balance after interest expenses
        if execute and stats["created"] > 0:
            try:
                api_patch(token, f"/items/accounts/{ACCOUNT_ID}", {
                    "balance": round(running_balance, 2)
                })
                print(f"  Updated account balance: K {account_balance:,.2f} → K {running_balance:,.2f}")
            except Exception as e:
                stats["errors"].append(f"Update account balance: {e}")

    return stats


# ─── Step 3: Process Other Expenses ─────────────────────────────────────────

def process_expenses(other_expenses, tracking_fees, token, execute=False):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    stats = {"created": 0, "errors": []}

    # Check what's already uploaded — expenses are stored in `transactions` as "Operational Cost"
    existing = api_get(token, "/items/transactions", {
        "filter[company][_eq]": COMPANY_ID,
        "filter[transaction_type][_eq]": "Operational Cost",
        "filter[transaction_date][_gte]": month_start_iso(),
        "filter[transaction_date][_lte]": month_end_iso(),
        "fields": "id,reference_number,notes,amount",
        "limit": -1,
    })
    existing_refs = {e.get("reference_number", "") for e in existing}
    existing_set = {
        (e.get("notes", "").lower(), round(float(e.get("amount", 0)), 0))
        for e in existing
    }
    print(f"  {len(existing)} existing {MONTH_NAME} expense transactions")

    # Combine tracking fees
    all_expenses = list(other_expenses)
    for tf in tracking_fees:
        all_expenses.append({
            "description": f"Tracking Fee - {tf['name']}",
            "amount": tf["fee"],
            "date": tf["date"],
        })

    exp_csv = os.path.join(OUTPUT_DIR, "expenses.csv")
    skipped = 0
    with open(exp_csv, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["date", "description", "amount", "reference", "action"])

        # Fetch account balance for running balance tracking
        account_balance = 0
        if execute:
            try:
                acct = api_get(token, f"/items/accounts/{ACCOUNT_ID}", {"fields": "id,balance"})
                if isinstance(acct, list):
                    acct = acct[0] if acct else {}
                account_balance = float(acct.get("balance", 0) or 0)
                print(f"  Account {ACCOUNT_ID} balance: K {account_balance:,.2f}")
            except Exception:
                pass
        running_balance = account_balance

        for idx, exp in enumerate(all_expenses):
            key = (exp["description"].lower(), round(exp["amount"], 0))
            if key in existing_set:
                writer.writerow([exp["date"], exp["description"], exp["amount"], "", "SKIP"])
                skipped += 1
                continue

            ref = f"EXP-{MONTH_ABBREV}-{idx+1:05d}"
            writer.writerow([exp["date"], exp["description"], exp["amount"], ref, "CREATE"])
            stats["created"] += 1

            if execute:
                running_balance -= exp["amount"]
                try:
                    api_post(token, "/items/transactions", {
                        "amount": exp["amount"],
                        "payment_amount": exp["amount"],
                        "transfer_fees": 0,
                        "transaction_date": exp["date"],
                        "payment_method": "cash",
                        "transaction_type": "Operational Cost",
                        "reference_number": ref,
                        "notes": exp["description"],
                        "transaction_status": "Completed",
                        "is_loan_transaction": False,
                        "is_debt_repayment": False,
                        "is_gps_fee_payment": False,
                        "new_amount": round(running_balance, 2),
                        "company": COMPANY_ID,
                        "account": ACCOUNT_ID,
                    })
                    if idx < 5:
                        print(f"  + Expense: {exp['description']} K{exp['amount']:,.2f}")
                    elif idx == 5:
                        print(f"  + ... uploading remaining expenses ...")
                    time.sleep(0.05)
                except Exception as e:
                    stats["errors"].append(f"Expense {exp['description']}: {e}")
            else:
                if idx < 3:
                    print(f"  [DRY] Would create: {exp['description']} K{exp['amount']:,.2f}")
                elif idx == 3:
                    remaining = stats["created"] - 3
                    print(f"  [DRY] ... and {remaining}+ more expenses")

        # Update account balance after all expenses
        if execute and stats["created"] > 0:
            try:
                api_patch(token, f"/items/accounts/{ACCOUNT_ID}", {
                    "balance": round(running_balance, 2)
                })
                print(f"  Updated account balance: K {account_balance:,.2f} → K {running_balance:,.2f}")
            except Exception as e:
                stats["errors"].append(f"Update account balance: {e}")

    if skipped:
        print(f"  Skipped {skipped} already-existing expenses")

    return stats


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Jutem Fund Monthly Delta Update")
    parser.add_argument("--sheet", required=True,
                        help="Excel sheet name to process, e.g. 'APRIL 2026'")
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK,
                        help=f"Path to workbook (default: {DEFAULT_WORKBOOK})")
    parser.add_argument("--execute", action="store_true",
                        help="Actually upload (default: dry run)")
    parser.add_argument("--step", choices=["loans", "interest", "expenses", "all"],
                        default="all", help="Which step to run")
    args = parser.parse_args()

    configure_run(args.sheet, args.workbook)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Parse Excel
    ws = load_excel()
    loans, interest_expenses, tracking_fees, other_expenses = parse_all_sections(ws)

    print(f"\n{'='*60}")
    print(f"{MONTH_NAME} Data Summary (from {os.path.basename(EXCEL_FILE)})")
    print(f"{'='*60}")
    print(f"  Loans:              {len(loans)}")
    paid = sum(1 for l in loans if l['status'] in ('paid', 'cleared'))
    pending = sum(1 for l in loans if l['status'] == 'pending')
    print(f"    - Paid:           {paid}")
    print(f"    - Pending:        {pending}")
    print(f"  Interest Expenses:  {len(interest_expenses)}")
    print(f"  Tracking Fees:      {len(tracking_fees)}")
    print(f"  Other Expenses:     {len(other_expenses)}")

    total_loaned = sum(l["amount"] for l in loans)
    total_expected = sum(l["expected_amount"] for l in loans)
    total_profit = sum(l["profit"] for l in loans)
    total_ie = sum(ie["interest_expense"] for ie in interest_expenses)
    total_expenses = sum(e["amount"] for e in other_expenses)
    total_tracking = sum(tf["fee"] for tf in tracking_fees)

    print(f"\n  Total Loaned:       K {total_loaned:>15,.2f}")
    print(f"  Total Expected:     K {total_expected:>15,.2f}")
    print(f"  Interest Income:    K {total_profit:>15,.2f}")
    print(f"  Interest Expense:   K {total_ie:>15,.2f}")
    print(f"  Net Interest:       K {total_profit - total_ie:>15,.2f}")
    print(f"  Other Expenses:     K {total_expenses:>15,.2f}")
    print(f"  Profit Before Tax:  K {total_profit - total_ie + total_tracking - total_expenses:>15,.2f}")
    print(f"{'='*60}")

    if not args.execute:
        print(f"\n  ** DRY RUN MODE — no changes will be made **\n")

    token = get_token()

    all_stats = {}

    if args.step in ("loans", "all"):
        print(f"\n{'─'*50}")
        print(f"Step 1: Loans, Amortization & Payments")
        print(f"{'─'*50}")
        all_stats["loans"] = process_loans(loans, token, execute=args.execute)

    if args.step in ("interest", "all"):
        print(f"\n{'─'*50}")
        print(f"Step 2: Interest Expenses")
        print(f"{'─'*50}")
        all_stats["interest"] = process_interest_expenses(interest_expenses, token, execute=args.execute)

    if args.step in ("expenses", "all"):
        print(f"\n{'─'*50}")
        print(f"Step 3: Other Expenses & Tracking Fees")
        print(f"{'─'*50}")
        all_stats["expenses"] = process_expenses(other_expenses, tracking_fees, token, execute=args.execute)

    # Summary
    print(f"\n{'='*60}")
    print("RESULTS SUMMARY")
    print(f"{'='*60}")

    if "loans" in all_stats:
        s = all_stats["loans"]
        print(f"\nLoans & Amortization:")
        print(f"  Matched existing amort:   {s['matched_existing']}")
        print(f"  Amort entries updated:    {s['amort_updated']} (expected_amount/profit/status)")
        print(f"  Statuses → paid:          {s['status_to_paid']}")
        print(f"  Payments created:         {s['payments_created']}")
        print(f"  Loan statuses updated:    {s['loans_status_updated']}")
        print(f"  New borrowers:            {s['new_borrowers']}")
        print(f"  New loans:                {s['new_loans']}")
        print(f"  New amortization entries: {s['new_amortization']}")
        if s['errors']:
            print(f"  Errors:                   {len(s['errors'])}")
            for err in s['errors'][:5]:
                print(f"    - {err}")

    if "interest" in all_stats:
        s = all_stats["interest"]
        print(f"\nInterest Expenses:")
        print(f"  Created:  {s['created']}")
        print(f"  Skipped:  {s['skipped']}")
        if s['errors']:
            print(f"  Errors:   {len(s['errors'])}")

    if "expenses" in all_stats:
        s = all_stats["expenses"]
        print(f"\nOther Expenses:")
        print(f"  Created:  {s['created']}")
        if s['errors']:
            print(f"  Errors:   {len(s['errors'])}")

    print(f"\nAudit CSVs in: {OUTPUT_DIR}/")

    if not args.execute:
        print(f"\nTo execute for real:")
        print(f"  export DIRECTUS_TOKEN='your-token'")
        print(f"  python3 etl/monthly_update.py --sheet {SHEET_NAME!r} --execute")

    summary = {
        "mode": "execute" if args.execute else "dry_run",
        "sheet": SHEET_NAME,
        "parsed": {
            "loans_total": len(loans),
            "loans_paid": paid,
            "loans_pending": pending,
            "interest_expenses": len(interest_expenses),
            "tracking_fees": len(tracking_fees),
            "other_expenses": len(other_expenses),
        },
        "totals": {
            "loans_amount": round(total_loaned, 2),
            "expected_amount": round(total_expected, 2),
            "interest_income": round(total_profit, 2),
            "interest_expense": round(total_ie, 2),
            "other_expenses": round(total_expenses, 2),
            "tracking_fees": round(total_tracking, 2),
        },
        "actions": {},
        "errors": [],
    }

    if "loans" in all_stats:
        s = all_stats["loans"]
        summary["actions"].update({
            "matched_existing": s["matched_existing"],
            "amort_updated": s["amort_updated"],
            "status_to_paid": s["status_to_paid"],
            "payments_created": s["payments_created"],
            "loans_status_updated": s["loans_status_updated"],
            "new_borrowers": s["new_borrowers"],
            "new_loans": s["new_loans"],
            "new_amortization": s["new_amortization"],
        })
        summary["errors"].extend(s["errors"])

    if "interest" in all_stats:
        s = all_stats["interest"]
        summary["actions"].update({
            "interest_expenses_created": s["created"],
            "interest_expenses_skipped": s["skipped"],
        })
        summary["errors"].extend(s["errors"])

    if "expenses" in all_stats:
        s = all_stats["expenses"]
        summary["actions"].update({
            "expenses_created": s["created"],
        })
        summary["errors"].extend(s["errors"])

    print()
    print("[SUMMARY]")
    print(json.dumps(summary))
    print("[/SUMMARY]")


if __name__ == "__main__":
    main()
